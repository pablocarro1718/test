"""
Genera un archivo Excel con el dashboard de inversiones completo.
Incluye todas las hojas, fórmulas y formato listo para usar.

Para ejecutar: python crear_excel_dashboard.py
"""

import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference, LineChart
from openpyxl.chart.label import DataLabelList

DIR = Path(__file__).parent

# Estilos
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14, color="2E75B6")
MONEY_FORMAT = '#,##0.00 €'
PERCENT_FORMAT = '0.00%'
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def aplicar_estilo_header(ws, row, cols):
    """Aplica estilo de cabecera a una fila."""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')


def ajustar_anchos(ws):
    """Ajusta el ancho de las columnas automáticamente."""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width


def crear_hoja_datos(wb):
    """Crea la hoja con los datos del CSV."""
    ws = wb.active
    ws.title = "Datos"

    # Leer CSV
    csv_path = DIR / "inversiones_unificadas.csv"
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader, 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Convertir números
                if row_idx > 1 and col_idx in [7, 8, 10, 11, 12, 13, 14]:  # columnas numéricas
                    try:
                        cell.value = float(value) if value else 0
                    except:
                        pass

    # Estilo de cabecera
    aplicar_estilo_header(ws, 1, 15)
    ajustar_anchos(ws)

    # Congelar primera fila
    ws.freeze_panes = 'A2'

    return ws


def crear_hoja_dashboard(wb):
    """Crea el dashboard con métricas principales."""
    ws = wb.create_sheet("Dashboard")

    # Título
    ws['A1'] = "DASHBOARD DE INVERSIONES"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:C1')

    # Sección: Métricas principales
    ws['A3'] = "MÉTRICAS PRINCIPALES"
    ws['A3'].font = Font(bold=True, size=12)

    metricas = [
        ("Total Invertido", '=ABS(SUMIFS(Datos!N:N,Datos!C:C,"BUY",Datos!A:A,"<>PENDIENTE"))', "Suma de todas las compras con fecha"),
        ("Total Recuperado", '=SUMIFS(Datos!N:N,Datos!C:C,"SELL")', "Suma de todas las ventas"),
        ("Dividendos Recibidos", '=SUMIFS(Datos!N:N,Datos!C:C,"DIVIDEND")', "Suma de dividendos"),
        ("Resultado Bruto", '=B6+B7-B5', "Recuperado + Dividendos - Invertido"),
        ("Comisiones Pagadas", '=SUMIF(Datos!M:M,">0",Datos!M:M)', "Total de comisiones"),
        ("Resultado Neto", '=B8-B9', "Resultado menos comisiones"),
        ("Rentabilidad", '=B10/B5', "Resultado / Inversión"),
    ]

    # Cabeceras
    ws['A4'] = "Métrica"
    ws['B4'] = "Valor"
    ws['C4'] = "Notas"
    aplicar_estilo_header(ws, 4, 3)

    for i, (nombre, formula, nota) in enumerate(metricas, 5):
        ws.cell(row=i, column=1, value=nombre)
        ws.cell(row=i, column=2, value=formula)
        ws.cell(row=i, column=3, value=nota)
        ws.cell(row=i, column=3).font = Font(italic=True, color="666666")

        # Formato moneda (excepto rentabilidad)
        if "Rentabilidad" not in nombre:
            ws.cell(row=i, column=2).number_format = MONEY_FORMAT
        else:
            ws.cell(row=i, column=2).number_format = PERCENT_FORMAT

    # Sección: Por Broker
    ws['A14'] = "RESUMEN POR BROKER"
    ws['A14'].font = Font(bold=True, size=12)

    ws['A15'] = "Broker"
    ws['B15'] = "Invertido"
    ws['C15'] = "Recuperado"
    ws['D15'] = "Resultado"
    aplicar_estilo_header(ws, 15, 4)

    brokers = ["Degiro", "Trading212", "Kraken"]
    for i, broker in enumerate(brokers, 16):
        ws.cell(row=i, column=1, value=broker)
        ws.cell(row=i, column=2, value=f'=ABS(SUMIFS(Datos!N:N,Datos!B:B,"{broker}",Datos!C:C,"BUY",Datos!A:A,"<>PENDIENTE"))')
        ws.cell(row=i, column=3, value=f'=SUMIFS(Datos!N:N,Datos!B:B,"{broker}",Datos!C:C,"SELL")+SUMIFS(Datos!N:N,Datos!B:B,"{broker}",Datos!C:C,"DIVIDEND")')
        ws.cell(row=i, column=4, value=f'=C{i}-B{i}')

        ws.cell(row=i, column=2).number_format = MONEY_FORMAT
        ws.cell(row=i, column=3).number_format = MONEY_FORMAT
        ws.cell(row=i, column=4).number_format = MONEY_FORMAT

    # Sección: Por Tipo de Activo
    ws['A21'] = "RESUMEN POR TIPO DE ACTIVO"
    ws['A21'].font = Font(bold=True, size=12)

    ws['A22'] = "Tipo"
    ws['B22'] = "Invertido"
    ws['C22'] = "Recuperado"
    ws['D22'] = "Resultado"
    aplicar_estilo_header(ws, 22, 4)

    tipos = ["Stock", "ETF", "Crypto"]
    for i, tipo in enumerate(tipos, 23):
        ws.cell(row=i, column=1, value=tipo)
        ws.cell(row=i, column=2, value=f'=ABS(SUMIFS(Datos!N:N,Datos!D:D,"{tipo}",Datos!C:C,"BUY",Datos!A:A,"<>PENDIENTE"))')
        ws.cell(row=i, column=3, value=f'=SUMIFS(Datos!N:N,Datos!D:D,"{tipo}",Datos!C:C,"SELL")+SUMIFS(Datos!N:N,Datos!D:D,"{tipo}",Datos!C:C,"DIVIDEND")')
        ws.cell(row=i, column=4, value=f'=C{i}-B{i}')

        ws.cell(row=i, column=2).number_format = MONEY_FORMAT
        ws.cell(row=i, column=3).number_format = MONEY_FORMAT
        ws.cell(row=i, column=4).number_format = MONEY_FORMAT

    # Ajustar anchos
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    return ws


def crear_hoja_crypto(wb):
    """Crea la hoja de posiciones abiertas de crypto."""
    ws = wb.create_sheet("Crypto_Abiertas")

    # Título
    ws['A1'] = "POSICIONES ABIERTAS - CRYPTO (KRAKEN)"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:G1')

    # Instrucción
    ws['A2'] = "Actualiza los precios en la columna D manualmente o usa =GOOGLEFINANCE(\"BTCEUR\")"
    ws['A2'].font = Font(italic=True, color="666666", size=10)
    ws.merge_cells('A2:G2')

    # Cabeceras
    headers = ["Ticker", "Cantidad", "Coste Total (€)", "Precio Actual (€)", "Valor Actual (€)", "P&L (€)", "P&L (%)"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    aplicar_estilo_header(ws, 4, 7)

    # Datos por crypto
    cryptos = [
        ("BTC", '=SUMIFS(Datos!G:G,Datos!E:E,"BTC",Datos!C:C,"BUY")', '=ABS(SUMIFS(Datos!N:N,Datos!E:E,"BTC",Datos!C:C,"BUY"))'),
        ("ETH", '=SUMIFS(Datos!G:G,Datos!E:E,"ETH",Datos!C:C,"BUY")', '=ABS(SUMIFS(Datos!N:N,Datos!E:E,"ETH",Datos!C:C,"BUY"))'),
        ("SOL", '=SUMIFS(Datos!G:G,Datos!E:E,"SOL",Datos!C:C,"BUY")', '=ABS(SUMIFS(Datos!N:N,Datos!E:E,"SOL",Datos!C:C,"BUY"))'),
    ]

    for i, (ticker, cant_formula, coste_formula) in enumerate(cryptos, 5):
        row = i
        ws.cell(row=row, column=1, value=ticker)
        ws.cell(row=row, column=2, value=cant_formula)  # Cantidad
        ws.cell(row=row, column=3, value=coste_formula)  # Coste
        ws.cell(row=row, column=4, value=0)  # Precio actual (manual)
        ws.cell(row=row, column=4).fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        ws.cell(row=row, column=5, value=f'=B{row}*D{row}')  # Valor actual
        ws.cell(row=row, column=6, value=f'=E{row}-C{row}')  # P&L €
        ws.cell(row=row, column=7, value=f'=IF(C{row}>0,F{row}/C{row},0)')  # P&L %

        # Formatos
        ws.cell(row=row, column=2).number_format = '0.00000000'
        ws.cell(row=row, column=3).number_format = MONEY_FORMAT
        ws.cell(row=row, column=4).number_format = MONEY_FORMAT
        ws.cell(row=row, column=5).number_format = MONEY_FORMAT
        ws.cell(row=row, column=6).number_format = MONEY_FORMAT
        ws.cell(row=row, column=7).number_format = PERCENT_FORMAT

    # Fila TOTAL
    ws.cell(row=8, column=1, value="TOTAL")
    ws.cell(row=8, column=1).font = Font(bold=True)
    ws.cell(row=8, column=3, value='=SUM(C5:C7)')
    ws.cell(row=8, column=5, value='=SUM(E5:E7)')
    ws.cell(row=8, column=6, value='=SUM(F5:F7)')
    ws.cell(row=8, column=7, value='=IF(C8>0,F8/C8,0)')

    ws.cell(row=8, column=3).number_format = MONEY_FORMAT
    ws.cell(row=8, column=5).number_format = MONEY_FORMAT
    ws.cell(row=8, column=6).number_format = MONEY_FORMAT
    ws.cell(row=8, column=7).number_format = PERCENT_FORMAT

    # Ajustar anchos
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 18

    return ws


def crear_hoja_operaciones_cerradas(wb):
    """Crea la hoja con análisis de operaciones cerradas."""
    ws = wb.create_sheet("Operaciones_Cerradas")

    # Título
    ws['A1'] = "ANÁLISIS DE OPERACIONES CERRADAS"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:F1')

    # Esta hoja es más compleja, añadimos instrucciones
    ws['A3'] = "Top 5 Mejores Operaciones (ordenar manualmente por P&L en hoja Datos)"
    ws['A3'].font = Font(bold=True, size=11)

    # Cabeceras
    headers = ["Ticker", "Fecha Compra", "Fecha Venta", "Invertido (€)", "Recuperado (€)", "P&L (€)"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    aplicar_estilo_header(ws, 4, 6)

    # Espacio para datos manuales o futura automatización
    ws['A5'] = "(Completa manualmente o filtra desde Datos)"
    ws['A5'].font = Font(italic=True, color="999999")

    # Resumen rápido
    ws['A12'] = "RESUMEN RÁPIDO"
    ws['A12'].font = Font(bold=True, size=12)

    ws['A13'] = "Operaciones de compra (con fecha)"
    ws['B13'] = '=COUNTIFS(Datos!C:C,"BUY",Datos!A:A,"<>PENDIENTE")'

    ws['A14'] = "Operaciones de venta"
    ws['B14'] = '=COUNTIF(Datos!C:C,"SELL")'

    ws['A15'] = "Dividendos recibidos"
    ws['B15'] = '=COUNTIF(Datos!C:C,"DIVIDEND")'

    ws['A16'] = "Operaciones pendientes de fecha"
    ws['B16'] = '=COUNTIF(Datos!A:A,"PENDIENTE")'

    # Ajustar anchos
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15

    return ws


def crear_hoja_tir(wb):
    """Crea la hoja para calcular la TIR."""
    ws = wb.create_sheet("Calcular_TIR")

    # Título
    ws['A1'] = "CÁLCULO DE TIR (XIRR)"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:C1')

    # Instrucciones
    instrucciones = [
        "Para calcular la TIR money-weighted necesitas:",
        "1. Copiar fechas y flujos de la hoja Datos (solo operaciones con fecha válida)",
        "2. Añadir una última fila con fecha=HOY y flujo=valor actual de crypto",
        "3. Usar la fórmula =XIRR(flujos, fechas)",
        "",
        "Ejemplo simplificado abajo:"
    ]

    for i, texto in enumerate(instrucciones, 3):
        ws.cell(row=i, column=1, value=texto)
        ws['A' + str(i)].font = Font(italic=True, color="666666")

    # Ejemplo
    ws['A10'] = "Fecha"
    ws['B10'] = "Flujo (€)"
    aplicar_estilo_header(ws, 10, 2)

    ws['A11'] = "2020-04-20"
    ws['B11'] = -1000  # Ejemplo inversión
    ws['A12'] = "2021-06-15"
    ws['B12'] = -500
    ws['A13'] = "2024-09-26"
    ws['B13'] = 2000  # Ejemplo recuperación
    ws['A14'] = "=TODAY()"
    ws['B14'] = 500  # Valor actual

    ws['A16'] = "TIR Anual:"
    ws['B16'] = '=XIRR(B11:B14,A11:A14)'
    ws['B16'].number_format = PERCENT_FORMAT
    ws['A16'].font = Font(bold=True)

    ws['A18'] = "NOTA: Este es un ejemplo. Debes reemplazar con tus datos reales."
    ws['A18'].font = Font(italic=True, color="FF0000")

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15

    return ws


def main():
    print("=" * 60)
    print("GENERANDO EXCEL CON DASHBOARD")
    print("=" * 60)

    wb = Workbook()

    print("\n[1/5] Creando hoja de Datos...")
    crear_hoja_datos(wb)

    print("[2/5] Creando Dashboard...")
    crear_hoja_dashboard(wb)

    print("[3/5] Creando hoja de Crypto...")
    crear_hoja_crypto(wb)

    print("[4/5] Creando hoja de Operaciones Cerradas...")
    crear_hoja_operaciones_cerradas(wb)

    print("[5/5] Creando hoja de TIR...")
    crear_hoja_tir(wb)

    # Guardar
    output_path = DIR / "portfolio_inversiones.xlsx"
    wb.save(output_path)

    print(f"\n✓ Archivo generado: {output_path}")
    print("\nPróximos pasos:")
    print("1. Descarga el archivo .xlsx de GitHub")
    print("2. Súbelo a Google Drive")
    print("3. Ábrelo con Google Sheets")
    print("4. Actualiza los precios de crypto en la hoja 'Crypto_Abiertas'")

    print("\n" + "=" * 60)


if __name__ == "__main__":
    main()
